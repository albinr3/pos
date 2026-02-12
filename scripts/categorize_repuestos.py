#!/usr/bin/env python3
"""Categoriza repuestos de un Excel usando OpenAI y agrega una columna `categoria`.

Uso:
  python scripts/categorize_repuestos.py --input repuestos.xlsx
  python scripts/categorize_repuestos.py --input repuestos.xlsx --output repuestos_categorizado.xlsx --batch-size 40
  python scripts/categorize_repuestos.py --input productos-20260211-2103.xlsx --carroceria-only
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib import error as urlerror
from urllib import request as urlrequest

from openpyxl import load_workbook
from tqdm import tqdm


CATEGORIES: List[str] = [
    "Filtros",
    "Frenos",
    "Luces",
    "Motor",
    "Refrigeración",
    "Rodamientos",
    "Sistema eléctrico",
    "Suspension y direccion",
    "Transmisión",
]
CARROCERIA_CATEGORY = "Carroceria"

MODEL_NAME = "gpt-5-mini"
OPENAI_CHAT_URL = "https://api.openai.com/v1/chat/completions"


def normalize_text(value: str) -> str:
    value = value.strip().lower()
    value = "".join(ch for ch in unicodedata.normalize("NFD", value) if unicodedata.category(ch) != "Mn")
    value = re.sub(r"\s+", " ", value)
    return value


def load_dotenv_if_needed(dotenv_path: Path) -> None:
    if os.getenv("OPENAI_API_KEY"):
        return

    if not dotenv_path.exists():
        return

    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def resolve_header_indices(headers: Sequence[Optional[str]]) -> Tuple[int, int, int]:
    normalized = {idx + 1: normalize_text(str(col or "")) for idx, col in enumerate(headers)}

    def find_first(candidates: Iterable[str]) -> Optional[int]:
        candidate_set = {normalize_text(x) for x in candidates}
        for idx, name in normalized.items():
            if name in candidate_set:
                return idx
        return None

    sku_idx = find_first(["sku", "codigo", "cod", "codigo producto", "code"])
    description_idx = find_first(
        [
            "f_descripcion",
            "descripcion",
            "descripcion corta",
            "descripcion completa",
            "description",
            "nombre",
            "producto",
        ]
    )
    reference_idx = find_first(["f_referencia_suplidor", "referencia", "reference", "ref"])

    missing = []
    if sku_idx is None:
        missing.append("sku")
    if description_idx is None:
        missing.append("descripcion")
    if reference_idx is None:
        missing.append("referencia")

    if missing:
        raise ValueError(
            f"No se encontraron columnas requeridas: {', '.join(missing)}. "
            f"Encabezados detectados: {[str(h or '') for h in headers]}"
        )

    return sku_idx, description_idx, reference_idx


def canonicalize_category(value: str) -> Optional[str]:
    if not value:
        return None
    norm = normalize_text(value)

    for category in CATEGORIES:
        if normalize_text(category) == norm:
            return category

    aliases = {
        "refrigeracion": "Refrigeración",
        "sistema electrico": "Sistema eléctrico",
        "suspension y direccion": "Suspension y direccion",
        "suspension/direccion": "Suspension y direccion",
        "transmision": "Transmisión",
    }
    return aliases.get(norm)


def parse_yes_no(value: str) -> Optional[bool]:
    if not value:
        return None
    norm = normalize_text(value)
    if norm in {"si", "sí", "yes", "true", "1", "carroceria", "carroceria si"}:
        return True
    if norm in {"no", "false", "0", "no carroceria"}:
        return False
    return None


def keyword_fallback(sku: str, description: str, reference: str) -> str:
    text = normalize_text(f"{sku} {description} {reference}")

    rules = [
        ("Filtros", ["filtro", "air filter", "oil filter", "fuel filter", "cabina"]),
        ("Frenos", ["freno", "pastilla", "disco", "balata", "caliper", "tambor"]),
        ("Luces", ["luz", "faro", "bombillo", "led", "halogena", "stop", "intermitente"]),
        ("Refrigeración", ["radiador", "coolant", "termostato", "ventilador", "anticongelante", "bomba de agua"]),
        ("Rodamientos", ["rodamiento", "bearing", "balinera", "cubo"]),
        ("Sistema eléctrico", ["alternador", "arranque", "bateria", "sensor", "rele", "fusible", "bobina"]),
        ("Suspension y direccion", ["amortiguador", "suspension", "direccion", "terminal", "cremallera", "buje", "rotula"]),
        ("Transmisión", ["clutch", "embrague", "transmision", "caja", "diferencial", "homocinetica"]),
    ]

    for category, words in rules:
        if any(word in text for word in words):
            return category

    return "Motor"


def keyword_fallback_carroceria(sku: str, description: str, reference: str) -> bool:
    text = normalize_text(f"{sku} {description} {reference}")
    carroceria_words = [
        "carroceria",
        "parachoque",
        "bumper",
        "guardalodo",
        "salpicadera",
        "capot",
        "bonete",
        "puerta",
        "compuerta",
        "tapa baul",
        "baul",
        "fender",
        "parrilla",
        "rejilla",
        "espejo",
        "retrovisor",
        "manija exterior",
        "bisagra puerta",
        "spoiler",
        "moldura",
    ]
    return any(word in text for word in carroceria_words)


def extract_json_from_text(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        match = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
        if match:
            return match.group(1).strip()
    return text


def call_openai_chat(api_key: str, prompt: str, max_completion_tokens: int = 2600) -> str:
    payload = {
        "model": MODEL_NAME,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Eres un clasificador de repuestos automotrices. "
                    "Debes devolver JSON valido y usar solo categorias permitidas."
                ),
            },
            {"role": "user", "content": prompt},
        ],
        "max_completion_tokens": max_completion_tokens,
    }

    req = urlrequest.Request(
        OPENAI_CHAT_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlrequest.urlopen(req, timeout=120) as response:
            body = response.read().decode("utf-8")
    except urlerror.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code}: {details[:800]}") from exc

    data = json.loads(body)
    choices = data.get("choices") or []
    if not choices:
        raise RuntimeError(f"Respuesta sin choices: {body[:500]}")

    choice0 = choices[0]
    content = choice0.get("message", {}).get("content")
    finish_reason = str(choice0.get("finish_reason", ""))
    if not content:
        if finish_reason == "length":
            raise RuntimeError(f"Respuesta truncada por limite de tokens (finish_reason=length): {body[:500]}")
        raise RuntimeError(f"Respuesta sin contenido: {body[:500]}")

    return content


def classify_batch(
    api_key: str,
    rows: Sequence[Dict[str, str]],
    retries: int,
    retry_base_sleep: float,
    max_completion_tokens: int,
) -> Dict[int, str]:
    categories_text = "\n".join(f"- {c}" for c in CATEGORIES)
    prompt = (
        "Clasifica cada producto en una sola categoria usando exclusivamente estas categorias:\n"
        f"{categories_text}\n\n"
        "Reglas:\n"
        "1) Responde SOLO JSON valido (sin markdown).\n"
        "2) Debes devolver exactamente un resultado por cada item.\n"
        "3) No inventes nuevas categorias.\n"
        "4) Si hay duda, elige la categoria mas probable por descripcion/referencia.\n\n"
        "Formato exacto de salida:\n"
        '{"items":[{"r":123,"c":"Motor"}]}\n\n'
        "Items a clasificar:\n"
        f"{json.dumps(rows, ensure_ascii=False)}"
    )

    last_error: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            raw = call_openai_chat(
                api_key=api_key,
                prompt=prompt,
                max_completion_tokens=max_completion_tokens,
            )
            parsed = json.loads(extract_json_from_text(raw))
            items = parsed.get("items")
            if not isinstance(items, list):
                raise ValueError("La respuesta no contiene 'items' como lista")

            result: Dict[int, str] = {}
            for item in items:
                if not isinstance(item, dict):
                    continue
                row = item.get("r", item.get("row"))
                category = canonicalize_category(str(item.get("c", item.get("categoria", ""))))
                if isinstance(row, int) and category:
                    result[row] = category

            return result
        except (json.JSONDecodeError, ValueError, RuntimeError, urlerror.HTTPError, urlerror.URLError) as exc:
            last_error = exc
            if attempt >= retries:
                break
            sleep_seconds = retry_base_sleep * attempt
            time.sleep(sleep_seconds)

    raise RuntimeError(f"No se pudo clasificar lote despues de {retries} intentos: {last_error}")


def classify_batch_carroceria(
    api_key: str,
    rows: Sequence[Dict[str, str]],
    retries: int,
    retry_base_sleep: float,
    max_completion_tokens: int,
) -> Dict[int, bool]:
    prompt = (
        "Determina si cada producto pertenece a la categoria Carroceria.\n"
        "Responde SOLO JSON valido (sin markdown), un item por fila.\n"
        "Formato exacto:\n"
        '{"items":[{"r":123,"es_carroceria":"SI"}]}\n'
        "Valores permitidos en es_carroceria: SI o NO.\n\n"
        "Items:\n"
        f"{json.dumps(rows, ensure_ascii=False)}"
    )

    last_error: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            raw = call_openai_chat(
                api_key=api_key,
                prompt=prompt,
                max_completion_tokens=max_completion_tokens,
            )
            parsed = json.loads(extract_json_from_text(raw))
            items = parsed.get("items")
            if not isinstance(items, list):
                raise ValueError("La respuesta no contiene 'items' como lista")

            result: Dict[int, bool] = {}
            for item in items:
                if not isinstance(item, dict):
                    continue
                row = item.get("r", item.get("row"))
                flag = parse_yes_no(str(item.get("es_carroceria", item.get("c", ""))))
                if isinstance(row, int) and flag is not None:
                    result[row] = flag
            return result
        except (json.JSONDecodeError, ValueError, RuntimeError, urlerror.HTTPError, urlerror.URLError) as exc:
            last_error = exc
            if attempt >= retries:
                break
            time.sleep(retry_base_sleep * attempt)

    raise RuntimeError(f"No se pudo clasificar lote de carroceria despues de {retries} intentos: {last_error}")


def classify_batch_resilient(
    api_key: str,
    rows: Sequence[Dict[str, str]],
    retries: int,
    retry_base_sleep: float,
    max_completion_tokens: int,
    progress: tqdm,
) -> Dict[int, str]:
    try:
        return classify_batch(
            api_key=api_key,
            rows=rows,
            retries=retries,
            retry_base_sleep=retry_base_sleep,
            max_completion_tokens=max_completion_tokens,
        )
    except Exception as exc:
        if len(rows) == 1:
            progress.write(
                f"Aviso: fila {rows[0].get('row')} no pudo clasificarse con IA; se usara fallback. Motivo: {exc}"
            )
            return {}

        midpoint = len(rows) // 2
        left = rows[:midpoint]
        right = rows[midpoint:]

        progress.write(
            f"Aviso: lote de {len(rows)} filas fallo. Se reintenta dividiendo en {len(left)} + {len(right)}."
        )

        out: Dict[int, str] = {}
        out.update(
            classify_batch_resilient(
                api_key=api_key,
                rows=left,
                retries=retries,
                retry_base_sleep=retry_base_sleep,
                max_completion_tokens=max_completion_tokens,
                progress=progress,
            )
        )
        out.update(
            classify_batch_resilient(
                api_key=api_key,
                rows=right,
                retries=retries,
                retry_base_sleep=retry_base_sleep,
                max_completion_tokens=max_completion_tokens,
                progress=progress,
            )
        )
        return out


def classify_batch_carroceria_resilient(
    api_key: str,
    rows: Sequence[Dict[str, str]],
    retries: int,
    retry_base_sleep: float,
    max_completion_tokens: int,
    progress: tqdm,
) -> Dict[int, bool]:
    try:
        return classify_batch_carroceria(
            api_key=api_key,
            rows=rows,
            retries=retries,
            retry_base_sleep=retry_base_sleep,
            max_completion_tokens=max_completion_tokens,
        )
    except Exception as exc:
        if len(rows) == 1:
            progress.write(
                f"Aviso: fila {rows[0].get('row')} no pudo clasificarse (Carroceria) con IA; se usara fallback. Motivo: {exc}"
            )
            return {}

        midpoint = len(rows) // 2
        left = rows[:midpoint]
        right = rows[midpoint:]
        progress.write(
            f"Aviso: lote de {len(rows)} filas (Carroceria) fallo. Se reintenta dividiendo en {len(left)} + {len(right)}."
        )

        out: Dict[int, bool] = {}
        out.update(
            classify_batch_carroceria_resilient(
                api_key=api_key,
                rows=left,
                retries=retries,
                retry_base_sleep=retry_base_sleep,
                max_completion_tokens=max_completion_tokens,
                progress=progress,
            )
        )
        out.update(
            classify_batch_carroceria_resilient(
                api_key=api_key,
                rows=right,
                retries=retries,
                retry_base_sleep=retry_base_sleep,
                max_completion_tokens=max_completion_tokens,
                progress=progress,
            )
        )
        return out


def process_excel(
    input_path: Path,
    output_path: Path,
    sheet_name: Optional[str],
    batch_size: int,
    retries: int,
    retry_base_sleep: float,
    max_completion_tokens: int,
    autosave_every_batches: int,
    limit: Optional[int],
) -> None:
    load_dotenv_if_needed(Path(".env"))
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("No se encontro OPENAI_API_KEY en el entorno ni en .env")

    source_path = output_path if output_path.exists() else input_path
    wb = load_workbook(filename=str(source_path))
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    sku_idx, description_idx, reference_idx = resolve_header_indices(headers)

    category_col = None
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        if normalize_text(str(cell_val or "")) == "categoria":
            category_col = col
            break

    if category_col is None:
        category_col = ws.max_column + 1
        ws.cell(row=1, column=category_col, value="categoria")

    rows_to_classify: List[Dict[str, str]] = []
    already_categorized = 0
    for row_number in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=row_number, column=sku_idx).value or "").strip()
        description = str(ws.cell(row=row_number, column=description_idx).value or "").strip()
        reference = str(ws.cell(row=row_number, column=reference_idx).value or "").strip()
        current_category = str(ws.cell(row=row_number, column=category_col).value or "").strip()

        if not sku and not description and not reference:
            ws.cell(row=row_number, column=category_col, value="")
            continue

        if current_category:
            already_categorized += 1
            continue

        rows_to_classify.append(
            {
                "row": row_number,
                "sku": sku,
                "descripcion": description,
                "referencia": reference,
            }
        )

    if limit is not None:
        rows_to_classify = rows_to_classify[:limit]

    if not rows_to_classify:
        wb.save(str(output_path))
        print("No hay filas pendientes por clasificar.")
        return

    missing_count = 0
    processed_batches = 0
    if source_path == output_path:
        print(f"Reanudando desde: {output_path}")
    if already_categorized:
        print(f"Filas ya categorizadas detectadas y omitidas: {already_categorized}")

    with tqdm(total=len(rows_to_classify), desc="Categorizando", unit="prod") as progress:
        for start in range(0, len(rows_to_classify), batch_size):
            batch = rows_to_classify[start : start + batch_size]
            classified = classify_batch_resilient(
                api_key=api_key,
                rows=batch,
                retries=retries,
                retry_base_sleep=retry_base_sleep,
                max_completion_tokens=max_completion_tokens,
                progress=progress,
            )

            for item in batch:
                row_number = int(item["row"])
                category = classified.get(row_number)
                if category is None:
                    missing_count += 1
                    category = keyword_fallback(
                        sku=item.get("sku", ""),
                        description=item.get("descripcion", ""),
                        reference=item.get("referencia", ""),
                    )

                ws.cell(row=row_number, column=category_col, value=category)
                progress.update(1)

            processed_batches += 1
            if autosave_every_batches > 0 and processed_batches % autosave_every_batches == 0:
                wb.save(str(output_path))
                progress.write(f"Progreso guardado: {progress.n}/{progress.total}")

    wb.save(str(output_path))

    if missing_count:
        print(
            f"Proceso completado. Archivo: {output_path}. "
            f"{missing_count} filas usaron fallback por respuesta incompleta."
        )
    else:
        print(f"Proceso completado. Archivo: {output_path}")


def process_excel_carroceria(
    input_path: Path,
    output_path: Path,
    sheet_name: Optional[str],
    batch_size: int,
    retries: int,
    retry_base_sleep: float,
    max_completion_tokens: int,
    autosave_every_batches: int,
    limit: Optional[int],
) -> None:
    load_dotenv_if_needed(Path(".env"))
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("No se encontro OPENAI_API_KEY en el entorno ni en .env")

    source_path = output_path if output_path.exists() else input_path
    wb = load_workbook(filename=str(source_path))
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    sku_idx, description_idx, reference_idx = resolve_header_indices(headers)

    flag_col = None
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        if normalize_text(str(cell_val or "")) == "es_carroceria":
            flag_col = col
            break

    if flag_col is None:
        flag_col = ws.max_column + 1
        ws.cell(row=1, column=flag_col, value="es_carroceria")

    rows_to_classify: List[Dict[str, str]] = []
    already_done = 0
    for row_number in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=row_number, column=sku_idx).value or "").strip()
        description = str(ws.cell(row=row_number, column=description_idx).value or "").strip()
        reference = str(ws.cell(row=row_number, column=reference_idx).value or "").strip()
        current_flag = str(ws.cell(row=row_number, column=flag_col).value or "").strip()

        if not sku and not description and not reference:
            ws.cell(row=row_number, column=flag_col, value="")
            continue

        if current_flag:
            already_done += 1
            continue

        rows_to_classify.append(
            {
                "row": row_number,
                "sku": sku,
                "descripcion": description,
                "referencia": reference,
            }
        )

    if limit is not None:
        rows_to_classify = rows_to_classify[:limit]

    if not rows_to_classify:
        wb.save(str(output_path))
        print("No hay filas pendientes por identificar en Carroceria.")
        return

    missing_count = 0
    processed_batches = 0
    if source_path == output_path:
        print(f"Reanudando desde: {output_path}")
    if already_done:
        print(f"Filas ya identificadas y omitidas: {already_done}")

    with tqdm(total=len(rows_to_classify), desc="Identificando Carroceria", unit="prod") as progress:
        for start in range(0, len(rows_to_classify), batch_size):
            batch = rows_to_classify[start : start + batch_size]
            classified = classify_batch_carroceria_resilient(
                api_key=api_key,
                rows=batch,
                retries=retries,
                retry_base_sleep=retry_base_sleep,
                max_completion_tokens=max_completion_tokens,
                progress=progress,
            )

            for item in batch:
                row_number = int(item["row"])
                flag = classified.get(row_number)
                if flag is None:
                    missing_count += 1
                    flag = keyword_fallback_carroceria(
                        sku=item.get("sku", ""),
                        description=item.get("descripcion", ""),
                        reference=item.get("referencia", ""),
                    )

                ws.cell(row=row_number, column=flag_col, value="SI" if flag else "NO")
                progress.update(1)

            processed_batches += 1
            if autosave_every_batches > 0 and processed_batches % autosave_every_batches == 0:
                wb.save(str(output_path))
                progress.write(f"Progreso guardado: {progress.n}/{progress.total}")

    wb.save(str(output_path))
    total_yes = 0
    for row_number in range(2, ws.max_row + 1):
        if normalize_text(str(ws.cell(row=row_number, column=flag_col).value or "")) == "si":
            total_yes += 1

    if missing_count:
        print(
            f"Proceso completado. Archivo: {output_path}. "
            f"{missing_count} filas usaron fallback. Coincidencias Carroceria: {total_yes}"
        )
    else:
        print(f"Proceso completado. Archivo: {output_path}. Coincidencias Carroceria: {total_yes}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Clasifica repuestos en un Excel y agrega columna categoria usando OpenAI"
    )
    parser.add_argument("--input", required=True, help="Ruta del Excel de entrada")
    parser.add_argument("--output", help="Ruta del Excel de salida")
    parser.add_argument("--sheet", help="Nombre de hoja (si no se indica, usa la hoja activa)")
    parser.add_argument(
        "--carroceria-only",
        action="store_true",
        help="Identifica solo si pertenece a Carroceria y escribe SI/NO en columna es_carroceria",
    )
    parser.add_argument("--batch-size", type=int, default=40, help="Cantidad de filas por llamada a IA")
    parser.add_argument("--retries", type=int, default=4, help="Reintentos por lote")
    parser.add_argument("--retry-base-sleep", type=float, default=1.5, help="Espera base entre reintentos")
    parser.add_argument(
        "--max-completion-tokens",
        type=int,
        default=2600,
        help="Maximo de tokens de salida por llamada (subir si hay truncamientos)",
    )
    parser.add_argument(
        "--autosave-every-batches",
        type=int,
        default=3,
        help="Guardar progreso cada N lotes (0 desactiva)",
    )
    parser.add_argument("--limit", type=int, help="Limita filas a procesar (util para pruebas)")
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"No existe el archivo de entrada: {input_path}", file=sys.stderr)
        return 1

    if args.output:
        output_path = Path(args.output)
    elif args.carroceria_only:
        output_path = input_path.with_name(f"{input_path.stem}_carroceria.xlsx")
    else:
        output_path = input_path.with_name(f"{input_path.stem}_categorizado.xlsx")

    if args.batch_size < 1:
        print("--batch-size debe ser >= 1", file=sys.stderr)
        return 1
    if args.max_completion_tokens < 200:
        print("--max-completion-tokens debe ser >= 200", file=sys.stderr)
        return 1
    if args.autosave_every_batches < 0:
        print("--autosave-every-batches debe ser >= 0", file=sys.stderr)
        return 1

    try:
        if args.carroceria_only:
            process_excel_carroceria(
                input_path=input_path,
                output_path=output_path,
                sheet_name=args.sheet,
                batch_size=args.batch_size,
                retries=args.retries,
                retry_base_sleep=args.retry_base_sleep,
                max_completion_tokens=args.max_completion_tokens,
                autosave_every_batches=args.autosave_every_batches,
                limit=args.limit,
            )
        else:
            process_excel(
                input_path=input_path,
                output_path=output_path,
                sheet_name=args.sheet,
                batch_size=args.batch_size,
                retries=args.retries,
                retry_base_sleep=args.retry_base_sleep,
                max_completion_tokens=args.max_completion_tokens,
                autosave_every_batches=args.autosave_every_batches,
                limit=args.limit,
            )
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
